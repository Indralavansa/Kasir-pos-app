import os
import sys
import shutil
from datetime import datetime, timedelta, timezone
from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, session, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, text
from sqlalchemy.orm import joinedload
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf import FlaskForm, CSRFProtect
from flask_wtf.csrf import CSRFError, generate_csrf
from wtforms import StringField, PasswordField, SubmitField, DecimalField, IntegerField, SelectField, TextAreaField, HiddenField
from wtforms.validators import DataRequired, Length, NumberRange, ValidationError, Optional
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from datetime import datetime as dt, date
import json
import io
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font
from pathlib import Path
from dotenv import load_dotenv

# Load environment variables from .env
load_dotenv()

print('=' * 60)
print('KASIR TOKO SEMBAKO - WITH SIMPLE BACKUP SYSTEM')
print('=' * 60)

# ==================== TIMEZONE HELPER ====================

def get_local_now():
    """Get current time in local timezone"""
    # Gunakan waktu lokal sistem (otomatis detect timezone dari OS)
    return datetime.now()

def get_local_timezone_name():
    """Get local timezone name (WIB/WITA/WIT/etc)"""
    import time
    
    # Get offset dari UTC dalam detik
    if time.daylight:
        offset_sec = -time.altzone
    else:
        offset_sec = -time.timezone
    
    # Convert ke jam
    offset_hours = offset_sec // 3600
    
    # Map ke nama timezone Indonesia
    timezone_map = {
        7: 'WIB',   # UTC+7 - Waktu Indonesia Barat
        8: 'WITA',  # UTC+8 - Waktu Indonesia Tengah
        9: 'WIT',   # UTC+9 - Waktu Indonesia Timur
    }
    
    # Return nama timezone atau offset format
    if offset_hours in timezone_map:
        return timezone_map[offset_hours]
    else:
        # Untuk timezone lain, tampilkan UTC offset
        sign = '+' if offset_hours >= 0 else '-'
        return f'UTC{sign}{abs(offset_hours)}'

# Debug: Print current time saat startup
local_tz_name = get_local_timezone_name()
print(f'[TIME] Local System Time: {get_local_now().strftime("%Y-%m-%d %H:%M:%S")}')
print(f'[TIME] Detected Timezone: {local_tz_name}')
print('=' * 60)

# ==================== SIMPLE BACKUP ====================

def backup_database():
    """Backup database dengan error handling yang lebih baik"""
    # Get base directory (parent of app folder)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    db_file = os.path.join(base_dir, 'instance', 'kasir.db')
    backup_folder = os.path.join(base_dir, 'backups')
    
    # Validasi path
    if not os.path.exists(os.path.dirname(db_file)):
        print(f"[Backup] ✗ Folder instance tidak ditemukan")
        return False
    
    # Pastikan folder backup ada
    try:
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder, exist_ok=True)
            print(f"[Backup] Created folder: {backup_folder}")
    except PermissionError:
        print(f"[Backup] ✗ Tidak ada izin membuat folder backup")
        return False
    
    if os.path.exists(db_file):
        try:
            # Cek ukuran database
            db_size = os.path.getsize(db_file)
            if db_size == 0:
                print("[Backup] ⚠️ Database kosong, backup dibatalkan")
                return False
            
            # Nama file backup
            timestamp = get_local_now().strftime('%Y%m%d_%H%M%S')
            backup_file = f'{backup_folder}/kasir_backup_{timestamp}.db'
            
            # Copy database dengan verifikasi
            shutil.copy2(db_file, backup_file)
            
            # Verifikasi backup
            if os.path.exists(backup_file):
                backup_size = os.path.getsize(backup_file)
                if backup_size == db_size:
                    print(f'[Backup] ✓ {backup_file}')
                    print(f'[Backup] Size: {backup_size/1024:.1f} KB')
                    
                    # Hapus backup lama (simpan 10 terbaru)
                    try:
                        backups = sorted([
                            f for f in os.listdir(backup_folder) 
                            if f.startswith('kasir_backup_') and f.endswith('.db')
                        ])
                        if len(backups) > 10:
                            for old_backup in backups[:-10]:
                                old_path = os.path.join(backup_folder, old_backup)
                                os.remove(old_path)
                                print(f'[Backup] Hapus backup lama: {old_backup}')
                    except Exception as e:
                        print(f'[Backup] ⚠️ Gagal hapus backup lama: {e}')
                    
                    return True
                else:
                    print(f'[Backup] ✗ Ukuran backup tidak sesuai')
                    os.remove(backup_file)  # Hapus backup yang rusak
                    return False
            else:
                print(f'[Backup] ✗ File backup tidak terbuat')
                return False
                
        except Exception as e:
            print(f'[Backup] ✗ Error: {e}')
            return False
    else:
        print(f'[Backup] ✗ Database not found: {db_file}')
        return False

# ==================== FLASK APP ====================

# Get base directory
base_app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Database Configuration with better error handling
DATABASE_URL = os.getenv('DATABASE_URL', '').strip()

print(f"[DB] DATABASE_URL env set: {bool(DATABASE_URL)}")

# If not set or empty, use SQLite (Local development)
if not DATABASE_URL or DATABASE_URL == 'None':
    db_path = os.path.join(base_app_dir, 'instance', 'kasir.db')
    DATABASE_URL = f'sqlite:///{db_path.replace(os.sep, "/")}'
    print(f"[DB] Using SQLite: {DATABASE_URL[:50]}...")
else:
    # Validate & convert postgres:// to postgresql:// for SQLAlchemy 1.4+
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    
    # Add explicit psycopg driver for PostgreSQL (required for psycopg 3.x)
    if DATABASE_URL.startswith('postgresql://'):
        DATABASE_URL = DATABASE_URL.replace('postgresql://', 'postgresql+psycopg://', 1)
    
    # Validate that URL looks reasonable
    if not ('postgresql+psycopg://' in DATABASE_URL or 'postgresql://' in DATABASE_URL or 'sqlite:///' in DATABASE_URL):
        print(f"[DB] WARNING: Invalid DATABASE_URL format")
        print(f"[DB] Got: {DATABASE_URL[:50]}...")
        # Fallback to SQLite
        db_path = os.path.join(base_app_dir, 'instance', 'kasir.db')
        DATABASE_URL = f'sqlite:///{db_path.replace(os.sep, "/")}'
        print(f"[DB] Falling back to SQLite")
    else:
        print(f"[DB] Using PostgreSQL: {DATABASE_URL.split('@')[0][:50]}...")

app = Flask(__name__)

# Apply ProxyFix middleware for Render proxy headers (HTTPS, remote addr, etc)
app.wsgi_app = ProxyFix(
    app.wsgi_app,
    x_for=1,        # Trust X-Forwarded-For
    x_proto=1,      # Trust X-Forwarded-Proto
    x_host=1,       # Trust X-Forwarded-Host
    x_port=1,       # Trust X-Forwarded-Port
    x_prefix=1      # Trust X-Forwarded-Prefix
)

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'rahasia-sangat-rahasia-123456-dev')
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['REMEMBER_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = True
app.config['REMEMBER_COOKIE_SECURE'] = True
# Add connection pool settings for production
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 10,
    'pool_recycle': 3600,
    'pool_pre_ping': True,
}

print(f"[DB] Final DATABASE_URI: {DATABASE_URL[:70]}...")

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu'
login_manager.login_message_category = 'info'
csrf = CSRFProtect(app)

# ==================== CUSTOM UNAUTHORIZED HANDLER ====================
@login_manager.unauthorized_handler
def unauthorized():
    """Handle unauthorized access cleanly without redirect loops"""
    return redirect(url_for('login'))

# ==================== AUTOMATIC DATABASE INITIALIZATION ====================
print("[DB] Initializing database tables...")
with app.app_context():
    try:
        db.create_all()
        print("[DB] ✅ Database tables initialized!")
    except Exception as e:
        print(f"[DB] ⚠️  Table creation warning: {e}")

# ==================== MEMBER CONFIG ====================

POINTS_PER_RUPIAH = 10000  # 1 point per Rp 10.000
LEVEL_RULES = [
    (0, 'Bronze'),
    (1000, 'Silver'),
    (5000, 'Gold'),
]

def get_member_level(points):
    level = 'Bronze'
    for min_points, name in LEVEL_RULES:
        if points >= min_points:
            level = name
    return level

def calculate_points_from_total(total_rupiah):
    if total_rupiah <= 0:
        return 0
    return int(total_rupiah // POINTS_PER_RUPIAH)

# ==================== TEMPLATE FILTERS ====================

@app.template_filter('format_datetime')
def format_datetime_filter(dt_obj):
    """Format datetime untuk display (no timezone conversion, langsung tampilkan)"""
    if dt_obj is None:
        return None
    # Langsung return datetime object as-is untuk di-format di template
    return dt_obj

@app.context_processor
def inject_globals():
    return {
        'current_time': get_local_now(),
        'timezone_name': get_local_timezone_name()
    }

@app.route('/img/<path:filename>')
def serve_img(filename):
    return send_from_directory(os.path.join(base_app_dir, 'img'), filename)

# ==================== REQUEST HANDLERS ====================

@app.before_request
def check_navigation():
    # Force HTTP for all requests (prevent HTTPS auto-upgrade)
    if request.headers.get('X-Forwarded-Proto') == 'https':
        return redirect(request.url.replace('https://', 'http://'), code=301)
    
    if current_user.is_authenticated:
        # If user tries to access login page while logged in, redirect to index
        # But allow register page for admin to add new users
        if request.endpoint == 'login':
            return redirect(url_for('index'))
    else:
        # If not logged in, only allow login/register and public assets
        if request.endpoint and request.endpoint not in ['login', 'register', 'static', 'serve_img']:
            return redirect(url_for('login'))

@app.after_request
def add_security_headers(response):
    # Relaxed CSP for development - allow more sources
    response.headers['Content-Security-Policy'] = "default-src 'self' 'unsafe-inline' 'unsafe-eval' data: blob:; script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://code.jquery.com https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com https://cdnjs.cloudflare.com; font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com; img-src 'self' data: https://*; connect-src 'self' https://*; object-src 'none'; base-uri 'self'; form-action 'self'"
    # Disable HSTS for development
    response.headers['Strict-Transport-Security'] = 'max-age=0'
    # Prevent caching for authenticated or protected pages to avoid back-button access after logout
    if request.endpoint and request.endpoint not in ['static', 'login']:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    nama = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='kasir')
    
    def set_password(self, password):
        """Set password dengan hashing yang lebih aman"""
        self.password_hash = generate_password_hash(
            password, 
            method='pbkdf2:sha256', 
            salt_length=16
        )
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    @staticmethod
    def validate_password_strength(password):
        """Validasi kekuatan password"""
        if len(password) < 8:
            return False, "Password minimal 8 karakter"
        if not any(c.isupper() for c in password):
            return False, "Password harus mengandung huruf besar"
        if not any(c.isdigit() for c in password):
            return False, "Password harus mengandung angka"
        return True, "Password valid"

class Kategori(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(100), nullable=False, unique=True)
    deskripsi = db.Column(db.Text)
    produk = db.relationship('Produk', backref='kategori_ref', lazy=True)

class Member(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(100), nullable=False)
    no_telp = db.Column(db.String(30))
    alamat = db.Column(db.Text)
    catatan = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=get_local_now)
    points = db.Column(db.Integer, default=0)
    total_spent = db.Column(db.Float, default=0)

    def get_level(self):
        return get_member_level(self.points or 0)

class Produk(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kode = db.Column(db.String(50), unique=True, nullable=False)
    nama = db.Column(db.String(200), nullable=False)
    deskripsi = db.Column(db.Text)
    harga_beli = db.Column(db.Float, nullable=False)
    harga_jual = db.Column(db.Float, nullable=False)
    stok = db.Column(db.Integer, default=0)
    kategori_id = db.Column(db.Integer, db.ForeignKey('kategori.id'))
    minimal_stok = db.Column(db.Integer, default=5)
    satuan = db.Column(db.String(20), default='pcs')
    harga_variasi = db.relationship('HargaVariasi', backref='produk', lazy=True, cascade='all, delete-orphan', order_by='HargaVariasi.min_qty')
    
    def get_harga_by_qty(self, qty):
        """Dapatkan harga berdasarkan quantity"""
        # Cek apakah ada harga variasi
        if self.harga_variasi:
            # Sort descending by min_qty untuk cek dari qty terbesar
            for variant in reversed(self.harga_variasi):
                if qty >= variant.min_qty:
                    return variant.harga
        # Default ke harga jual
        return self.harga_jual
    
    def to_dict(self):
        # Get price variants
        variants = []
        if self.harga_variasi:
            variants = [{'min_qty': v.min_qty, 'harga': v.harga} for v in self.harga_variasi]
            
        return {
            'id': self.id,
            'kode': self.kode,
            'nama': self.nama,
            'harga_jual': self.harga_jual,
            'harga_variasi': variants,
            'stok': self.stok,
            'satuan': self.satuan
        }

class HargaVariasi(db.Model):
    """Model untuk menyimpan harga bertingkat berdasarkan quantity"""
    id = db.Column(db.Integer, primary_key=True)
    produk_id = db.Column(db.Integer, db.ForeignKey('produk.id'), nullable=False)
    min_qty = db.Column(db.Integer, nullable=False)  # Minimal quantity untuk harga ini
    harga = db.Column(db.Float, nullable=False)  # Harga per unit
    keterangan = db.Column(db.String(100))  # Opsional: keterangan tier harga

class Transaksi(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kode_transaksi = db.Column(db.String(50), unique=True, nullable=False)
    tanggal = db.Column(db.DateTime, default=get_local_now)
    subtotal = db.Column(db.Float, nullable=False, default=0)
    discount_percent = db.Column(db.Float, default=0)
    discount_amount = db.Column(db.Float, default=0)
    total = db.Column(db.Float, nullable=False)
    bayar = db.Column(db.Float, nullable=False)
    kembalian = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(20), default='tunai')  # tunai, qris, ewallet, debit, hutang
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'))
    member_manual = db.Column(db.String(100))  # Untuk input manual nama/telp member
    points_earned = db.Column(db.Integer, default=0)
    user = db.relationship('User', backref='transaksi')
    member = db.relationship('Member', backref='transaksi')
    items = db.relationship('TransaksiItem', backref='transaksi_ref', lazy=True)

class TransaksiItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    transaksi_id = db.Column(db.Integer, db.ForeignKey('transaksi.id'))
    produk_id = db.Column(db.Integer, db.ForeignKey('produk.id'))
    produk = db.relationship('Produk')
    jumlah = db.Column(db.Integer, nullable=False)
    harga = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)

class Pengaturan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.Text)
    
    @staticmethod
    def get(key, default=''):
        setting = Pengaturan.query.filter_by(key=key).first()
        return setting.value if setting else default
    
    @staticmethod
    def set(key, value):
        setting = Pengaturan.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            setting = Pengaturan(key=key, value=value)
            db.session.add(setting)
        db.session.commit()

# ==================== FORMS ====================

class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

class RegisterForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=3, max=80)])
    nama = StringField('Nama Lengkap', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Konfirmasi Password', validators=[DataRequired()])
    role = SelectField('Role', choices=[('kasir', 'Kasir'), ('admin', 'Admin')])
    submit = SubmitField('Register')
    
    def validate_username(self, username):
        """Validasi username"""
        username_val = username.data
        if ' ' in username_val:
            raise ValidationError('Username tidak boleh mengandung spasi')
        if len(username_val) < 3:
            raise ValidationError('Username minimal 3 karakter')
        
        user = User.query.filter_by(username=username_val).first()
        if user:
            raise ValidationError('Username sudah digunakan.')
    
    def validate_password(self, field):
        """Validasi kekuatan password"""
        password = field.data
        if len(password) < 8:
            raise ValidationError('Password minimal 8 karakter')
        if not any(c.isupper() for c in password):
            raise ValidationError('Password harus mengandung huruf besar')
        if not any(c.isdigit() for c in password):
            raise ValidationError('Password harus mengandung angka')
    
    def validate_confirm_password(self, confirm_password):
        if self.password.data != confirm_password.data:
            raise ValidationError('Password tidak cocok.')

class ProdukForm(FlaskForm):
    product_id = HiddenField()
    kode = StringField('Kode Produk', validators=[DataRequired()])
    nama = StringField('Nama Produk', validators=[DataRequired()])
    deskripsi = TextAreaField('Deskripsi')
    harga_beli = DecimalField('Harga Beli', validators=[DataRequired(), NumberRange(min=0)])
    harga_jual = DecimalField('Harga Jual', validators=[DataRequired(), NumberRange(min=0)])
    stok = IntegerField('Stok Awal', validators=[DataRequired(), NumberRange(min=0)])
    kategori_id = SelectField('Kategori', coerce=int)
    minimal_stok = IntegerField('Minimal Stok', validators=[DataRequired(), NumberRange(min=0)])
    satuan = StringField('Satuan', validators=[DataRequired()])
    submit = SubmitField('Simpan')
    
    def validate_kode(self, kode):
        produk = Produk.query.filter_by(kode=kode.data).first()
        if produk:
            # Allow same kode when editing the current product
            if self.product_id.data:
                try:
                    if produk.id == int(self.product_id.data):
                        return
                except ValueError:
                    pass
            raise ValidationError('Kode produk sudah digunakan! Gunakan kode yang berbeda.')

class KategoriForm(FlaskForm):
    nama = StringField('Nama Kategori', validators=[DataRequired()])
    deskripsi = TextAreaField('Deskripsi')
    submit = SubmitField('Simpan')

class MemberForm(FlaskForm):
    nama = StringField('Nama Member', validators=[DataRequired(), Length(max=100)])
    no_telp = StringField('Nomor Telepon', validators=[Optional(), Length(max=30)])
    alamat = TextAreaField('Alamat', validators=[Optional(), Length(max=500)])
    catatan = TextAreaField('Catatan', validators=[Optional(), Length(max=500)])
    submit = SubmitField('Simpan')

# ==================== USER LOADER ====================

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# ==================== ROUTES ====================

@app.route('/')
def index():
    # Manual redirect untuk menghindari loop dengan proxy Render
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    
    total_produk = Produk.query.count()
    today = date.today()
    total_transaksi_hari_ini = Transaksi.query.filter(db.func.date(Transaksi.tanggal) == today).count()
    produk_habis = Produk.query.filter(Produk.stok <= Produk.minimal_stok).count()
    
    return render_template('index.html', 
                         total_produk=total_produk,
                         total_transaksi=total_transaksi_hari_ini,
                         produk_habis=produk_habis,
                         current_time=get_local_now(),
                         timezone_name=get_local_timezone_name())

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Ensure admin user exists (first-time setup)
    try:
        admin_count = User.query.filter_by(username='admin').count()
        if admin_count == 0:
            admin = User(
                username='admin',
                email='admin@kasirtokosembako.local',
                is_active=True
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("[INIT] Admin user created: admin / admin123")
    except Exception as e:
        print(f"[INIT] Could not check/create admin user: {e}")
    
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = LoginForm()
    # Ensure CSRF token is generated and stored in session for the login form.
    generate_csrf()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            flash('Login berhasil!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Username atau password salah!', 'danger')
    
    return render_template('auth/login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('index'))
    
    form = RegisterForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            nama=form.nama.data,
            role=form.role.data
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash(f'User {form.username.data} berhasil ditambahkan!', 'success')
        return redirect(url_for('register'))
    
    # Get all users to display in table
    users = User.query.order_by(User.id).all()
    
    return render_template('auth/register.html', form=form, users=users)

@app.route('/user/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'}), 403
    
    user = User.query.get_or_404(user_id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        return jsonify({'success': False, 'message': 'Tidak bisa menghapus akun sendiri!'}), 400
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'User {username} berhasil dihapus!'})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'info')
    response = redirect(url_for('login'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# ==================== PRODUK ROUTES ====================

@app.route('/produk')
@login_required
def list_produk():
    filter_stok = request.args.get('filter', 'semua')
    search = request.args.get('q', '').strip()
    kategori_id = request.args.get('kategori_id', '').strip()
    
    query = Produk.query
    
    if filter_stok == 'habis':
        query = query.filter(Produk.stok == 0)
    elif filter_stok == 'hampir_habis':
        query = query.filter(Produk.stok > 0, Produk.stok <= Produk.minimal_stok)
    elif filter_stok == 'tersedia':
        query = query.filter(Produk.stok > 0)

    kategori_filter = None
    if kategori_id.isdigit():
        kategori_filter = Kategori.query.get(int(kategori_id))
        if kategori_filter:
            query = query.filter(Produk.kategori_id == kategori_filter.id)

    if search:
        query = query.filter(
            or_(
                Produk.nama.ilike(f'%{search}%'),
                Produk.kode.ilike(f'%{search}%')
            )
        )
    
    produk_list = query.order_by(Produk.kode).all()
    kategori_list = Kategori.query.all()
    return render_template('produk/list.html', 
                         produk_list=produk_list,
                         kategori_list=kategori_list,
                         filter_stok=filter_stok,
                         search=search,
                         kategori_filter=kategori_filter)

@app.route('/produk/tambah', methods=['GET', 'POST'])
@login_required
def tambah_produk():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_produk'))
    
    form = ProdukForm()
    form.kategori_id.choices = [(k.id, k.nama) for k in Kategori.query.all()]
    
    if form.validate_on_submit():
        try:
            produk = Produk(
                kode=form.kode.data,
                nama=form.nama.data,
                deskripsi=form.deskripsi.data,
                harga_beli=float(form.harga_beli.data),
                harga_jual=float(form.harga_jual.data),
                stok=form.stok.data,
                kategori_id=form.kategori_id.data if form.kategori_id.data else None,
                minimal_stok=form.minimal_stok.data,
                satuan=form.satuan.data
            )
            db.session.add(produk)
            db.session.flush()  # Get produk.id
            
            # Handle harga variasi
            variant_min_qtys = request.form.getlist('variant_min_qty[]')
            variant_hargas = request.form.getlist('variant_harga[]')
            variant_keterangans = request.form.getlist('variant_keterangan[]')
            
            for i in range(len(variant_min_qtys)):
                if variant_min_qtys[i] and variant_hargas[i]:
                    harga_variasi = HargaVariasi(
                        produk_id=produk.id,
                        min_qty=int(variant_min_qtys[i]),
                        harga=float(variant_hargas[i]),
                        keterangan=variant_keterangans[i] if i < len(variant_keterangans) else None
                    )
                    db.session.add(harga_variasi)
            
            db.session.commit()
            flash('Produk berhasil ditambahkan!', 'success')
            return redirect(url_for('list_produk'))
        except ValueError as e:
            db.session.rollback()
            flash(f'Error: Input tidak valid - {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menyimpan produk: {str(e)}', 'danger')
    
    return render_template('produk/form.html', form=form, title='Tambah Produk', produk=None)

@app.route('/produk/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_produk(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_produk'))
    
    produk = Produk.query.get_or_404(id)
    form = ProdukForm(obj=produk)
    form.product_id.data = produk.id
    form.kategori_id.choices = [(k.id, k.nama) for k in Kategori.query.all()]
    
    if form.validate_on_submit():
        try:
            produk.kode = form.kode.data
            produk.nama = form.nama.data
            produk.deskripsi = form.deskripsi.data
            produk.harga_beli = float(form.harga_beli.data)
            produk.harga_jual = float(form.harga_jual.data)
            produk.stok = form.stok.data
            produk.kategori_id = form.kategori_id.data if form.kategori_id.data else None
            produk.minimal_stok = form.minimal_stok.data
            produk.satuan = form.satuan.data
            
            # Update harga variasi - hapus yang lama, insert yang baru
            # Karena relationship cascade delete-orphan, akan auto-delete
            produk.harga_variasi.clear()
            
            variant_min_qtys = request.form.getlist('variant_min_qty[]')
            variant_hargas = request.form.getlist('variant_harga[]')
            variant_keterangans = request.form.getlist('variant_keterangan[]')
            
            for i in range(len(variant_min_qtys)):
                if variant_min_qtys[i] and variant_hargas[i]:
                    harga_variasi = HargaVariasi(
                        produk_id=produk.id,
                        min_qty=int(variant_min_qtys[i]),
                        harga=float(variant_hargas[i]),
                        keterangan=variant_keterangans[i] if i < len(variant_keterangans) else None
                    )
                    db.session.add(harga_variasi)
            
            db.session.commit()
            flash('Produk berhasil diperbarui!', 'success')
            return redirect(url_for('list_produk'))
        except ValueError as e:
            db.session.rollback()
            flash(f'Error: Input tidak valid - {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menyimpan produk: {str(e)}', 'danger')
    
    return render_template('produk/form.html', form=form, title='Edit Produk', produk=produk)

@app.route('/produk/hapus/<int:id>', methods=['POST'])
@login_required
def hapus_produk(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_produk'))
    
    produk = Produk.query.get_or_404(id)
    db.session.delete(produk)
    db.session.commit()
    flash('Produk berhasil dihapus!', 'success')
    return redirect(url_for('list_produk'))

# ==================== KATEGORI ROUTES ====================

@app.route('/kategori')
@login_required
def list_kategori():
    kategori_list = Kategori.query.all()
    return render_template('produk/kategori.html', kategori_list=kategori_list)

@app.route('/kategori/tambah', methods=['GET', 'POST'])
@login_required
def tambah_kategori():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_kategori'))
    
    form = KategoriForm()
    if form.validate_on_submit():
        try:
            kategori = Kategori(nama=form.nama.data, deskripsi=form.deskripsi.data)
            db.session.add(kategori)
            db.session.commit()
            flash('Kategori berhasil ditambahkan!', 'success')
            return redirect(url_for('list_kategori'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menambahkan kategori: {str(e)}', 'danger')
    
    return render_template('produk/kategori_form.html', form=form, title='Tambah Kategori')

@app.route('/kategori/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_kategori(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_kategori'))
    
    kategori = Kategori.query.get_or_404(id)
    form = KategoriForm(obj=kategori)
    
    if form.validate_on_submit():
        try:
            kategori.nama = form.nama.data
            kategori.deskripsi = form.deskripsi.data
            db.session.commit()
            flash('Kategori berhasil diperbarui!', 'success')
            return redirect(url_for('list_kategori'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui kategori: {str(e)}', 'danger')
    
    return render_template('produk/kategori_form.html', form=form, title='Edit Kategori')

@app.route('/kategori/hapus/<int:id>', methods=['POST'])
@login_required
def hapus_kategori(id):
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('list_kategori'))
    
    kategori = Kategori.query.get_or_404(id)
    db.session.delete(kategori)
    db.session.commit()
    flash('Kategori berhasil dihapus!', 'success')
    return redirect(url_for('list_kategori'))

@app.route('/kategori/<int:id>/produk')
@login_required
def kategori_produk(id):
    kategori = Kategori.query.get_or_404(id)
    produk_list = (Produk.query
        .filter(Produk.kategori_id == kategori.id)
        .options(joinedload(Produk.harga_variasi))
        .order_by(Produk.kode)
        .all())
    return render_template('produk/kategori_produk.html',
                         kategori=kategori,
                         produk_list=produk_list)

# ==================== MEMBER ROUTES ====================

@app.route('/member')
@login_required
def list_member():
    search = request.args.get('q', '').strip()
    query = Member.query

    if search:
        query = query.filter(
            or_(
                Member.nama.ilike(f'%{search}%'),
                Member.no_telp.ilike(f'%{search}%')
            )
        )

    member_list = query.order_by(Member.nama).all()
    return render_template('member/list.html', member_list=member_list, search=search)

@app.route('/member/<int:id>/transaksi')
@login_required
def member_transaksi(id):
    member = Member.query.get_or_404(id)
    transaksi_list = (
        Transaksi.query
        .filter_by(member_id=member.id)
        .order_by(Transaksi.tanggal.desc())
        .all()
    )
    total_transaksi = len(transaksi_list)
    total_belanja = sum(t.total for t in transaksi_list)
    return render_template(
        'member/transaksi.html',
        member=member,
        transaksi_list=transaksi_list,
        total_transaksi=total_transaksi,
        total_belanja=total_belanja
    )

@app.route('/member/export')
@login_required
def export_member():
    members = Member.query.order_by(Member.nama).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    headers = ['nama', 'no_telp', 'alamat', 'catatan', 'points', 'total_spent']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for member in members:
        ws.append([
            member.nama,
            member.no_telp or '',
            member.alamat or '',
            member.catatan or '',
            member.points or 0,
            member.total_spent or 0
        ])

    # Ranking sheet
    ws_rank = wb.create_sheet('Ranking')
    ws_rank.append(['Member', 'Total Spent'])
    for cell in ws_rank[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    top_members = sorted(members, key=lambda m: m.total_spent or 0, reverse=True)[:10]
    for member in top_members:
        ws_rank.append([member.nama, member.total_spent or 0])

    if top_members:
        data = Reference(ws_rank, min_col=2, min_row=1, max_row=len(top_members) + 1)
        categories = Reference(ws_rank, min_col=1, min_row=2, max_row=len(top_members) + 1)
        chart = BarChart()
        chart.title = 'Top Member by Total Spent'
        chart.y_axis.title = 'Total Spent'
        chart.x_axis.title = 'Member'
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 16
        ws_rank.add_chart(chart, 'D2')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = 'attachment; filename=member_export.xlsx'
    return response

@app.route('/member/template')
@login_required
def download_member_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    headers = ['nama', 'no_telp', 'alamat', 'catatan', 'points', 'total_spent']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    ws.append(['Siti Aminah', '081234567890', 'Jl. Melati No. 1', 'Pelanggan loyal', 0, 0])
    ws.append(['Budi Santoso', '082233445566', 'Jl. Kenanga No. 12', 'Suka belanja grosir', 120, 450000])
    ws.append(['Rina', '', 'Pasar Blok A', 'Tanpa no telp', 30, 125000])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = 'attachment; filename=member_template.xlsx'
    return response

@app.route('/member/import', methods=['POST'])
@login_required
def import_member():
    if 'file' not in request.files:
        flash('File XLSX tidak ditemukan.', 'danger')
        return redirect(url_for('list_member'))

    file = request.files['file']
    if not file or file.filename == '':
        flash('File XLSX tidak ditemukan.', 'danger')
        return redirect(url_for('list_member'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb['Members'] if 'Members' in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash('File XLSX kosong.', 'danger')
            return redirect(url_for('list_member'))

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        required_columns = ['nama', 'no_telp', 'alamat', 'catatan', 'points', 'total_spent']
        if not set(required_columns).issubset(set(headers)):
            flash('Header XLSX tidak sesuai. Gunakan template yang disediakan.', 'danger')
            return redirect(url_for('list_member'))

        col_index = {name: headers.index(name) for name in required_columns}
        added = 0
        updated = 0

        for row in rows[1:]:
            nama = (row[col_index['nama']] or '').strip() if row[col_index['nama']] else ''
            if not nama:
                continue

            no_telp_val = row[col_index['no_telp']]
            alamat_val = row[col_index['alamat']]
            catatan_val = row[col_index['catatan']]

            no_telp = str(no_telp_val).strip() if no_telp_val is not None else None
            alamat = str(alamat_val).strip() if alamat_val is not None else None
            catatan = str(catatan_val).strip() if catatan_val is not None else None
            no_telp = no_telp or None
            alamat = alamat or None
            catatan = catatan or None

            try:
                points = int(float(row[col_index['points']] or 0))
            except (TypeError, ValueError):
                points = 0

            try:
                total_spent = float(row[col_index['total_spent']] or 0)
            except (TypeError, ValueError):
                total_spent = 0

            if no_telp:
                existing = Member.query.filter_by(no_telp=no_telp).first()
            else:
                existing = Member.query.filter_by(nama=nama).first()
            if existing:
                existing.alamat = alamat
                existing.catatan = catatan
                existing.points = points
                existing.total_spent = total_spent
                updated += 1
            else:
                member = Member(
                    nama=nama,
                    no_telp=no_telp,
                    alamat=alamat,
                    catatan=catatan,
                    points=points,
                    total_spent=total_spent
                )
                db.session.add(member)
                added += 1

        db.session.commit()
        flash(f'Import selesai: {added} ditambah, {updated} diperbarui.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal import member: {str(e)}', 'danger')

    return redirect(url_for('list_member'))

@app.route('/member/tambah', methods=['GET', 'POST'])
@login_required
def tambah_member():
    form = MemberForm()

    if form.validate_on_submit():
        try:
            member = Member(
                nama=form.nama.data.strip(),
                no_telp=form.no_telp.data.strip() if form.no_telp.data else None,
                alamat=form.alamat.data.strip() if form.alamat.data else None,
                catatan=form.catatan.data.strip() if form.catatan.data else None
            )
            db.session.add(member)
            db.session.commit()
            flash('Member berhasil ditambahkan!', 'success')
            return redirect(url_for('list_member'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menambahkan member: {str(e)}', 'danger')

    return render_template('member/form.html', form=form, title='Tambah Member')

@app.route('/member/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_member(id):
    member = Member.query.get_or_404(id)
    form = MemberForm(obj=member)

    if form.validate_on_submit():
        try:
            member.nama = form.nama.data.strip()
            member.no_telp = form.no_telp.data.strip() if form.no_telp.data else None
            member.alamat = form.alamat.data.strip() if form.alamat.data else None
            member.catatan = form.catatan.data.strip() if form.catatan.data else None
            db.session.commit()
            flash('Member berhasil diperbarui!', 'success')
            return redirect(url_for('list_member'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui member: {str(e)}', 'danger')

    return render_template('member/form.html', form=form, title='Edit Member')

@app.route('/member/hapus/<int:id>', methods=['POST'])
@login_required
def hapus_member(id):
    member = Member.query.get_or_404(id)
    try:
        Transaksi.query.filter_by(member_id=member.id).update(
            {Transaksi.member_id: None},
            synchronize_session=False
        )
        db.session.delete(member)
        db.session.commit()
        flash('Member berhasil dihapus!', 'success')
        return redirect(url_for('list_member'))
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus member: {str(e)}', 'danger')
        return redirect(url_for('list_member'))

# ==================== TRANSAKSI ROUTES ====================

@app.route('/kasir')
@login_required
def kasir():
    kategori_list = Kategori.query.all()
    member_list = Member.query.order_by(Member.nama).all()
    # Eager load harga_variasi untuk performa dan avoid lazy loading issues
    produk_list = Produk.query.filter(Produk.stok > 0).options(db.joinedload(Produk.harga_variasi)).all()
    return render_template('transaksi/kasir.html', 
                         kategori_list=kategori_list,
                         produk_list=produk_list,
                         member_list=member_list)

@app.route('/api/produk')
@login_required
def get_api_produk():
    search = request.args.get('search', '').strip()
    kategori_id = request.args.get('kategori_id', type=int)
    
    query = Produk.query.filter(Produk.stok > 0)
    
    if search:
        query = query.filter(
            or_(
                Produk.nama.ilike(f'%{search}%'),
                Produk.kode.ilike(f'%{search}%')
            )
        )
    
    if kategori_id:
        query = query.filter_by(kategori_id=kategori_id)
    
    produk_list = query.limit(50).all()  # Batasi hasil untuk performa
    
    return jsonify([p.to_dict() for p in produk_list])

@app.route('/transaksi/checkout', methods=['POST'])
@login_required
@csrf.exempt
def checkout():
    print("\n" + "="*50)
    print("CHECKOUT PROCESS STARTED")
    print("="*50)
    
    try:
        # Log user dan IP
        user_info = f"User: {current_user.username} (ID: {current_user.id})"
        ip_info = f"IP: {request.remote_addr}"
        print(f"[Checkout] {user_info} | {ip_info}")
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Data tidak valid'})
        
        items = data.get('items', [])
        total = data.get('total', 0)
        bayar = data.get('bayar', 0)
        payment_method = data.get('payment_method', 'tunai')
        member_id = data.get('member_id')
        member_manual = data.get('member_manual')  # Input manual nama/telp
        
        print(f"[Checkout] Items: {len(items)}")
        print(f"[Checkout] Total: Rp {total}")
        print(f"[Checkout] Bayar: Rp {bayar}")
        
        if not items:
            return jsonify({'success': False, 'message': 'Keranjang kosong'})
        
        # Validasi tipe data
        try:
            total = float(total)
            bayar = float(bayar)
        except ValueError:
            return jsonify({'success': False, 'message': 'Format angka tidak valid'})
        
        member = None
        if member_id:
            try:
                member = Member.query.get(int(member_id))
            except (TypeError, ValueError):
                return jsonify({'success': False, 'message': 'Member tidak valid'})
            if not member:
                return jsonify({'success': False, 'message': 'Member tidak ditemukan'})

        # Calculate totals server-side
        subtotal = 0
        for item in items:
            try:
                qty = int(item.get('quantity', 1))
                price = float(item.get('price', 0))
            except (TypeError, ValueError):
                return jsonify({'success': False, 'message': 'Format item tidak valid'})
            if qty < 1 or price < 0:
                return jsonify({'success': False, 'message': 'Data item tidak valid'})
            subtotal += qty * price

        total = subtotal
        points_earned = calculate_points_from_total(total) if member else 0

        if bayar < total:
            return jsonify({'success': False, 'message': 'Pembayaran kurang'})
        
        kode_transaksi = f'TRX{get_local_now().strftime("%Y%m%d%H%M%S")}'
        print(f"[Checkout] Transaction code: {kode_transaksi}")
        print(f"[Checkout] Local Time: {get_local_now().strftime('%Y-%m-%d %H:%M:%S')} {get_local_timezone_name()}")
        
        # Buat transaksi
        transaksi = Transaksi(
            kode_transaksi=kode_transaksi,
            subtotal=subtotal,
            discount_percent=0,
            discount_amount=0,
            total=total,
            bayar=bayar,
            kembalian=bayar - total,
            payment_method=payment_method,
            user_id=current_user.id,
            member_id=member.id if member else None,
            member_manual=member_manual if not member else None,  # Simpan input manual jika bukan member terdaftar
            points_earned=points_earned
        )
        db.session.add(transaksi)
        db.session.flush()
        
        # Log timestamp transaksi
        print(f"[Checkout] Transaction timestamp: {transaksi.tanggal.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Tambahkan items
        for item in items:
            produk = Produk.query.get(item['id'])
            if not produk:
                return jsonify({'success': False, 'message': f'Produk tidak ditemukan'})
                
            quantity = item.get('quantity', 1)
            
            print(f"[Checkout] Item: {produk.nama} x{quantity}")
            
            if produk.stok < quantity:
                return jsonify({'success': False, 'message': f'Stok {produk.nama} tidak cukup'})
            
            transaksi_item = TransaksiItem(
                transaksi_id=transaksi.id,
                produk_id=item['id'],
                jumlah=quantity,
                harga=item['price'],
                subtotal=item['price'] * quantity
            )
            db.session.add(transaksi_item)
            produk.stok -= quantity
        
        # Commit transaksi
        db.session.commit()

        if member:
            member.points = (member.points or 0) + points_earned
            member.total_spent = (member.total_spent or 0) + total
            db.session.commit()
        print("[Checkout] ✓ Transaction saved to database")
        
        # === BACKUP SETELAH TRANSAKSI ===
        print("\n[Backup] Starting backup after transaction...")
        backup_success = backup_database()
        
        if backup_success:
            print("[Backup] ✓ Backup completed successfully!")
        else:
            print("[Backup] ✗ Backup failed!")
        
        print("="*50)
        print("CHECKOUT PROCESS COMPLETED")
        print("="*50 + "\n")
        
        return jsonify({
            'success': True,
            'kode_transaksi': kode_transaksi,
            'kembalian': bayar - total,
            'transaksi_id': transaksi.id,
            'total': total,
            'subtotal': subtotal,
            'discount_percent': 0,
            'discount_amount': 0,
            'points_earned': points_earned,
            'bayar': bayar,
            'payment_method': payment_method,
            'tanggal': transaksi.tanggal.strftime('%Y-%m-%d %H:%M:%S'),
            'kasir': current_user.nama
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"[Checkout] ✗ Error: {e}")
        import traceback
        error_detail = traceback.format_exc()
        print(f"[Checkout TRACEBACK] {error_detail}")
        
        return jsonify({
            'success': False, 
            'message': 'Terjadi kesalahan sistem'
        })

@app.route('/transaksi')
@login_required
def list_transaksi():
    # Get filter parameters
    tanggal_mulai = request.args.get('tanggal_mulai', '')
    tanggal_selesai = request.args.get('tanggal_selesai', '')
    payment_method = request.args.get('payment_method', '')
    kode = request.args.get('kode', '')
    
    query = Transaksi.query
    
    # Apply filters
    if tanggal_mulai:
        query = query.filter(db.func.date(Transaksi.tanggal) >= tanggal_mulai)
    if tanggal_selesai:
        query = query.filter(db.func.date(Transaksi.tanggal) <= tanggal_selesai)
    if payment_method:
        query = query.filter(Transaksi.payment_method == payment_method)
    if kode:
        query = query.filter(Transaksi.kode_transaksi.like(f'%{kode}%'))
    
    transaksi_list = query.order_by(Transaksi.tanggal.desc()).all()
    
    return render_template('transaksi/list.html', 
                         transaksi_list=transaksi_list,
                         tanggal_mulai=tanggal_mulai,
                         tanggal_selesai=tanggal_selesai,
                         payment_method=payment_method,
                         kode=kode,
                         timezone_name=get_local_timezone_name())

@app.route('/transaksi/struk/<int:id>')
@login_required
def view_struk(id):
    transaksi = Transaksi.query.get_or_404(id)
    settings = {
        'store_name': Pengaturan.get('store_name', 'TOKO SEMBAKO'),
        'store_address': Pengaturan.get('store_address', 'Jl. Contoh No. 123'),
        'store_phone': Pengaturan.get('store_phone', '021-12345678'),
        'receipt_footer': Pengaturan.get('receipt_footer', 'Terima kasih atas kunjungan Anda')
    }
    return render_template('transaksi/struk.html', 
                         transaksi=transaksi, 
                         settings=settings,
                         timezone_name=get_local_timezone_name())

@app.route('/laporan')
@login_required
def laporan():
    if current_user.role != 'admin':
        flash('Akses ditolak! Hanya admin yang bisa melihat laporan.', 'danger')
        return redirect(url_for('index'))
    
    tanggal_mulai = request.args.get('tanggal_mulai', date.today().strftime('%Y-%m-%d'))
    tanggal_selesai = request.args.get('tanggal_selesai', date.today().strftime('%Y-%m-%d'))
    
    transaksi_list = Transaksi.query.filter(
        db.func.date(Transaksi.tanggal) >= tanggal_mulai,
        db.func.date(Transaksi.tanggal) <= tanggal_selesai
    ).order_by(Transaksi.tanggal).all()
    
    total_penjualan = sum(t.total for t in transaksi_list)
    total_transaksi = len(transaksi_list)
    total_keuntungan = 0
    for t in transaksi_list:
        for item in t.items:
            harga_beli = item.produk.harga_beli if item.produk else 0
            total_keuntungan += (item.harga - harga_beli) * item.jumlah
    
    # Chart data: Sales per day
    from collections import defaultdict
    sales_by_date = defaultdict(float)
    for t in transaksi_list:
        date_str = t.tanggal.strftime('%Y-%m-%d')
        sales_by_date[date_str] += t.total
    
    # Chart data: Payment method distribution
    payment_count = defaultdict(int)
    for t in transaksi_list:
        payment_count[t.payment_method or 'tunai'] += 1
    
    # Chart data: Top 5 products
    from collections import Counter
    product_sales = Counter()
    for t in transaksi_list:
        for item in t.items:
            product_sales[item.produk.nama] += item.subtotal
    top_products = product_sales.most_common(5)

    top_members_rows = (db.session.query(
        Member.id,
        Member.nama,
        db.func.sum(Transaksi.total).label('total_spent'),
        db.func.count(Transaksi.id).label('total_transaksi')
    )
        .join(Transaksi, Transaksi.member_id == Member.id)
        .filter(
            db.func.date(Transaksi.tanggal) >= tanggal_mulai,
            db.func.date(Transaksi.tanggal) <= tanggal_selesai
        )
        .group_by(Member.id, Member.nama)
        .order_by(db.func.sum(Transaksi.total).desc())
        .limit(25)
        .all())

    top_members = [
        {
            'nama': row.nama,
            'total_spent': row.total_spent or 0,
            'total_transaksi': row.total_transaksi or 0
        }
        for row in top_members_rows
    ]
    
    return render_template('laporan/index.html',
                         transaksi_list=transaksi_list,
                         tanggal_mulai=tanggal_mulai,
                         tanggal_selesai=tanggal_selesai,
                         total_penjualan=total_penjualan,
                         total_keuntungan=total_keuntungan,
                         total_transaksi=total_transaksi,
                         sales_by_date=dict(sales_by_date),
                         payment_count=dict(payment_count),
                         top_products=top_products,
                         top_members=top_members,
                         timezone_name=get_local_timezone_name())

# ==================== PENGATURAN ROUTES ====================

@app.route('/pengaturan', methods=['GET', 'POST'])
@login_required
def pengaturan():
    if current_user.role != 'admin':
        flash('Akses ditolak! Hanya admin yang bisa mengakses pengaturan.', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Update settings
        Pengaturan.set('store_name', request.form.get('store_name', ''))
        Pengaturan.set('store_address', request.form.get('store_address', ''))
        Pengaturan.set('store_phone', request.form.get('store_phone', ''))
        Pengaturan.set('receipt_footer', request.form.get('receipt_footer', ''))
        Pengaturan.set('tax_enabled', request.form.get('tax_enabled', 'false'))
        Pengaturan.set('tax_percentage', request.form.get('tax_percentage', '0'))
        
        flash('Pengaturan berhasil disimpan!', 'success')
        return redirect(url_for('pengaturan'))
    
    # Get current settings
    settings = {
        'store_name': Pengaturan.get('store_name', 'TOKO SEMBAKO'),
        'store_address': Pengaturan.get('store_address', 'Jl. Contoh No. 123'),
        'store_phone': Pengaturan.get('store_phone', '021-12345678'),
        'receipt_footer': Pengaturan.get('receipt_footer', 'Terima kasih atas kunjungan Anda'),
        'tax_enabled': Pengaturan.get('tax_enabled', 'false'),
        'tax_percentage': Pengaturan.get('tax_percentage', '10')
    }
    
    return render_template('pengaturan/index.html', settings=settings)

# ==================== BACKUP ROUTES ====================

@app.route('/admin/backup-now')
@login_required
def backup_now():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('index'))
    
    if backup_database():
        flash('Backup berhasil dibuat!', 'success')
    else:
        flash('Gagal membuat backup!', 'danger')
    
    return redirect(url_for('index'))

@app.route('/admin/restore-backup')
@login_required
def restore_backup():
    if current_user.role != 'admin':
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('index'))
    
    # Implementasi restore nanti
    flash('Fitur restore dalam pengembangan', 'info')
    return redirect(url_for('index'))

# ==================== ERROR HANDLERS ====================

@app.errorhandler(400)
def bad_request(e):
    return render_template('errors/400.html'), 400

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403

@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('errors/500.html'), 500

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    print(f"[CSRF] {e.description}")
    return render_template('errors/400.html'), 400

# ==================== INIT DATABASE ====================

def init_database():
    with app.app_context():
        db.create_all()

        # Lightweight migration for existing databases
        try:
            result = db.session.execute(text("PRAGMA table_info(transaksi)"))
            tx_columns = {row[1] for row in result}
            if 'member_id' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN member_id INTEGER"))
            if 'subtotal' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN subtotal REAL DEFAULT 0"))
            if 'discount_percent' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN discount_percent REAL DEFAULT 0"))
            if 'discount_amount' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN discount_amount REAL DEFAULT 0"))
            if 'points_earned' not in tx_columns:
                db.session.execute(text("ALTER TABLE transaksi ADD COLUMN points_earned INTEGER DEFAULT 0"))

            result = db.session.execute(text("PRAGMA table_info(member)"))
            member_columns = {row[1] for row in result}
            if 'points' not in member_columns:
                db.session.execute(text("ALTER TABLE member ADD COLUMN points INTEGER DEFAULT 0"))
            if 'total_spent' not in member_columns:
                db.session.execute(text("ALTER TABLE member ADD COLUMN total_spent REAL DEFAULT 0"))

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[DB] Warning: gagal migrasi kolom member_id: {e}")
        
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', nama='Administrator', role='admin')
            admin.set_password('Admin123')
            db.session.add(admin)
            
            kasir = User(username='kasir', nama='Kasir Toko', role='kasir')
            kasir.set_password('Kasir123')
            db.session.add(kasir)
            
            # Kategori
            kategori_list = [
                Kategori(nama='Sembako', deskripsi='Bahan pokok sehari-hari'),
                Kategori(nama='Minuman', deskripsi='Minuman berbagai jenis'),
                Kategori(nama='Snack', deskripsi='Makanan ringan'),
                Kategori(nama='Bahan Masak', deskripsi='Bahan-bahan untuk memasak'),
                Kategori(nama='Lainnya', deskripsi='Produk lainnya'),
            ]
            for kategori in kategori_list:
                if not Kategori.query.filter_by(nama=kategori.nama).first():
                    db.session.add(kategori)
            
            # Produk sample
            if not Produk.query.first():
                produk_sample = [
                    ('BRG001', 'Beras Premium 5kg', 1, 45000, 50000, 50),
                    ('BRG002', 'Minyak Goreng 2L', 4, 25000, 28000, 30),
                    ('BRG003', 'Gula Pasir 1kg', 1, 12000, 14000, 40),
                    ('BRG004', 'Aqua Gelas 240ml', 2, 500, 1000, 100),
                    ('BRG005', 'Indomie Goreng', 3, 2500, 3000, 60),
                ]
                
                for kode, nama, kategori_id, harga_beli, harga_jual, stok in produk_sample:
                    produk = Produk(
                        kode=kode,
                        nama=nama,
                        kategori_id=kategori_id,
                        harga_beli=harga_beli,
                        harga_jual=harga_jual,
                        stok=stok,
                        minimal_stok=5,
                        satuan='pcs'
                    )
                    db.session.add(produk)
            
            db.session.commit()
            print('✓ Database initialized')
            print('  Admin: admin / Admin123')
            print('  Kasir: kasir / Kasir123')
            print('  Note: Password harus mengandung huruf besar dan angka')
            
            # Backup pertama
            print('\n[Backup] Creating first backup...')
            backup_database()

# ==================== RUN APP ====================

if __name__ == '__main__':
    init_database()
    print('\n' + '='*60)
    print('SERVER READY!')
    print('URL Lokal: http://localhost:5000')
    print('URL Jaringan: http://0.0.0.0:5000')
    print('Backup folder: backups/')
    print('='*60 + '\n')
    app.run(host='0.0.0.0', debug=True, port=5000)